import sys
import pandas as pd
import win32com.client as win32
import pythoncom
import openpyxl
from openpyxl import Workbook
import os
from AnyQt.QtWidgets import QApplication, QPushButton, QCheckBox
from Orange.widgets.settings import Setting
from Orange.widgets.utils.signals import Input, Output
from Orange.data import Domain, StringVariable, Table, DiscreteVariable

if "site-packages/Orange/widgets" in os.path.dirname(os.path.abspath(__file__)).replace("\\", "/"):
    from Orange.widgets.orangecontrib.AAIT.utils.thread_management import Thread
    from Orange.widgets.orangecontrib.AAIT.utils import  base_widget
    from Orange.widgets.orangecontrib.AAIT.utils.color_prefix_utils import normalize_hex, parse_color_prefix, hex_rgb_to_excel_bgr
else:
    from orangecontrib.AAIT.utils.thread_management import Thread
    from orangecontrib.AAIT.utils import  base_widget
    from orangecontrib.AAIT.utils.color_prefix_utils import normalize_hex, parse_color_prefix, hex_rgb_to_excel_bgr


class OWInsertDataToExcel(base_widget.BaseListWidget):
    name = "Insert Data to Excel"
    description = "Insère une data table dans un fichier Excel à partir d'une cellule définie."
    category = "AAIT - TOOLBOX"
    icon = "icons/xlsx_insert.png"
    if "site-packages/Orange/widgets" in os.path.dirname(os.path.abspath(__file__)).replace("\\", "/"):
        icon = "icons_dev/xlsx_insert.png"
    gui = os.path.join(os.path.dirname(__file__), "designer", "owinsertdatatoexcel.ui")
    want_control_area = False
    priority = 1010

    auto_send = Setting(True)
    selected_column_name = Setting("path")
    create_sheet_if_missing = Setting(False)
    create_file_if_missing = Setting(False)
    include_headers = Setting(False)


    def _normalize_hex(self, color: str) -> str:
        return normalize_hex(color)

    def parse_color_prefix(self, value: str):
        return parse_color_prefix(value)

    def _hex_rgb_to_excel_bgr(self, hex_rgb: str) -> int:
        return hex_rgb_to_excel_bgr(hex_rgb)

    class Inputs:
        parameters = Input("Parameters", Table)
        data       = Input("Data Table", Table)

    class Outputs:
        status_data = Output("Status Table", Table)
        output = Output("Output Table", Table)

    # ------------------------------------------------------------------ inputs
    @Inputs.parameters
    def set_parameters(self, in_data: Table | None):
        #A modif pour changer file_path en selection
        self.parameters = in_data
        if self.parameters:
            self.var_selector.add_variables(self.parameters.domain)
            self.var_selector.select_variable_by_name(self.selected_column_name)
        if self.auto_send and self.parameters and self.data:
            self.run()

    @Inputs.data
    def set_data(self, in_data: Table | None):
        self.data = in_data
        if self.auto_send and self.parameters and self.data:
            self.run()

    # ------------------------------------------------------------------ init
    def __init__(self):
        super().__init__()
        self.setSizeGripEnabled(False)
        self.setFixedSize(474, 490)
        self.parameters = None
        self.data       = None
        self.thread     = None
        self.result = None

        self.pushButton_run = self.findChild(QPushButton,  "pushButton_run")
        self.checkBox_send  = self.findChild(QCheckBox,    "checkBox_send")
        self.checkBox_create_sheet = self.findChild(QCheckBox, "checkBox_create_sheet")
        self.checkBox_create_missing_file = self.findChild(QCheckBox, "checkBox_create_missing_file")


        if self.checkBox_send:
            self.checkBox_send.setChecked(self.auto_send)
            self.checkBox_send.toggled.connect(self._on_autorun_toggled)

        if self.checkBox_create_sheet:
            self.checkBox_create_sheet.setChecked(self.create_sheet_if_missing)
            self.checkBox_create_sheet.toggled.connect(
                lambda v: setattr(self, "create_sheet_if_missing", v)
            )  

        if self.checkBox_create_missing_file:
            self.checkBox_create_missing_file.setChecked(self.create_file_if_missing)
            self.checkBox_create_missing_file.toggled.connect(
                lambda v: setattr(self, "create_file_if_missing", v)
            )

        self.checkBox_include_headers = self.findChild(QCheckBox, "checkBox_include_headers")

        if self.checkBox_include_headers:
            self.checkBox_include_headers.setChecked(self.include_headers)
            self.checkBox_include_headers.toggled.connect(
                lambda v: setattr(self, "include_headers", v)
            )

        if self.pushButton_run:
            self.pushButton_run.clicked.connect(self.run)


    # ------------------------------------------------------------------ slots UI
    def _on_autorun_toggled(self, checked: bool):
        self.auto_send = checked

    # ------------------------------------------------------------------ run
    def run(self):
        self.error("")
        self.warning("")
        self.information("")
        # If thread is running, quit previous
        if self.thread is not None:
            # Note: win32com threads are tricky to kill, but we follow the pattern
            if hasattr(self.thread, 'safe_quit'):
                self.thread.safe_quit()

        if self.parameters is None or self.data is None:
            return

        col_name = self.selected_column_name
        try:
            self.parameters.domain[col_name]
        except (KeyError, ValueError):
            self.error(f"Colonne '{col_name}' introuvable.")
            return

        try:            
            # On convertit en DataFrame pour faciliter le mapping des autres colonnes (sheet, row, col)
            params_df = self._orange_table_to_dataframe(self.parameters)
            params_df.columns = params_df.columns.str.strip().str.lower()

            data_df = self._orange_table_to_dataframe(self.data)

        except Exception as e:
            self.error(f"Erreur initialisation : {e}")
            return

        if self.pushButton_run:
            self.pushButton_run.setEnabled(False)
        # Start progress bar
        self.progressBarInit()

        # Connect and start thread
        self.thread = Thread(self._run_logic, params_df, data_df, self.create_sheet_if_missing, self.include_headers)
        self.thread.progress.connect(self.handle_progress)
        self.thread.result.connect(self.handle_result)
        self.thread.finish.connect(self.handle_finish)
        self.thread.start()

    def handle_progress(self, value: float) -> None:
        self.progressBarSet(value)

    def handle_result(self, results_list):
        if not results_list:
            return
        
        # On crée une table de statut globale
        domain = Domain([], metas=[
            StringVariable("file"), 
            DiscreteVariable("status", values=["ok", "ko"]), 
            StringVariable("details")
        ])
        
        table_data = [[r["file"], r["status"], r["details"]] for r in results_list]
        self.Outputs.status_data.send(Table.from_list(domain, table_data))
        # Notification sommaire
        errors = [r for r in results_list if r["status"] == "ko"]
        if errors:
            self.error(f"{len(errors)} erreur(s) détectée(s).")
            self.Outputs.output.send(None)
        else:
            self.information("Toutes les insertions ont réussi.")
            if self.data is not None:
                self.Outputs.output.send(self.data )
        self.parameters = None
        self.data = None

    def handle_finish(self):
        if self.pushButton_run:
            self.pushButton_run.setEnabled(True)
        self.progressBarFinished()
        print("Excel insertion process finished")

    # --- Thread Logic ---
    def _run_logic_com(self, params_df, df, create_sheet, include_headers, progress_callback):
        results = []
        excel = None
        try:
            pythoncom.CoInitialize()
            excel = win32.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False

            total_files = len(params_df)
            col_name = self.selected_column_name.strip().lower()

            for i, row in params_df.iterrows():
                file_path = str(row.get(col_name, "")).strip()
                # On prépare une liste pour collecter les logs de ce fichier précis
                logs = [] 

                file_params = {
                    "file_path": file_path,
                    "sheet_name": str(row.get("sheet_name", "Sheet1")).strip(),
                    "start_col": self.safe_int(row.get("col", 1), 1),
                    "start_row": self.safe_int(row.get("row", 1), 1)
                }

                try:
                    if not self.create_file_if_missing: # on ne se met en erreur que si on n autorise pas la creation
                        if not os.path.isfile(file_params["file_path"]):
                            results.append({
                                "file": file_params["file_path"],
                                "status": "ko",
                                "details": "Fichier introuvable",
                            })
                            continue

                        wb = excel.Workbooks.Open(os.path.abspath(file_params["file_path"]))
                    else:
                        abs_path = os.path.abspath(file_params["file_path"])
                        # si le dossier n existe pas je le crée
                        os.makedirs(os.path.dirname(abs_path), exist_ok=True)
                        # Création du fichier Excel s'il n'existe pas
                        if not os.path.isfile(abs_path):
                            logs.append("Fichier Excel introuvable, création automatique.")

                            # Crée un nouveau workbook
                            wb_new = excel.Workbooks.Add()

                            # Sauvegarde au chemin demandé
                            wb_new.SaveAs(abs_path)

                            wb_new.Close(False)

                        # Ouverture du fichier
                        wb = excel.Workbooks.Open(abs_path)

                    sheet_name = file_params["sheet_name"]

                    # --- Gestion intelligente des feuilles avec log interne ---
                    ws = None
                    try:
                        ws = wb.Worksheets(sheet_name)

                    except Exception:
                        logs.append(f"Feuille '{sheet_name}' introuvable.")

                        if create_sheet:
                            ws = wb.Worksheets.Add()
                            ws.Name = sheet_name
                            logs.append(f"Feuille '{sheet_name}' créée.")
                        else:
                            try:
                                ws = wb.Worksheets("Sheet1")
                                logs.append("Fallback sur 'Sheet1'.")
                            except Exception:
                                ws = wb.Worksheets(1)
                                logs.append("Fallback sur la première feuille.")

                    # 1. Préparation des données (on nettoie les balises pour l'affichage final)
                    if include_headers:
                        headers = [list(df.columns)]
                        data_matrix = headers + df.values.tolist()
                    else:
                        data_matrix = df.values.tolist()

                    # 2. Insertion des données en bloc
                    start_r = file_params["start_row"]
                    start_c = file_params["start_col"]

                    last_r = start_r + len(data_matrix) - 1
                    last_c = start_c + len(df.columns) - 1

                    target_range = ws.Range(ws.Cells(start_r, start_c), ws.Cells(last_r, last_c))
                    target_range.Value = data_matrix

                    # 3. Application des couleurs cellule par cellule si un préfixe %!color!% est détecté
                    header_offset = 1 if include_headers else 0
                    for r_idx, row_data in enumerate(data_matrix[header_offset:], start=header_offset):
                        for c_idx, cell_value in enumerate(row_data):
                            clean_value, hex_bg = self.parse_color_prefix(str(cell_value))

                            if hex_bg:
                                target_cell = ws.Cells(start_r + r_idx, start_c + c_idx)
                                target_cell.Interior.Color = self._hex_rgb_to_excel_bgr(hex_bg)
                                target_cell.Font.Color = 0x000000
                                target_cell.Value = clean_value

                    wb.Save()
                    wb.Close(False)

                    # On combine le succès avec les éventuels avertissements de fallback
                    detail_msg = "Insertion réussie"
                    if logs:
                        detail_msg += " (Note: " + " | ".join(logs) + ")"

                    results.append({
                        "file": file_params["file_path"],
                        "status": "ok",
                        "details": detail_msg
                    })

                except Exception as e:
                    results.append({
                        "file": file_params.get("file_path", "unknown"),
                        "status": "ko",
                        "details": str(e)
                    })

                progress_callback(int(((i + 1) / total_files) * 100))

            return results

        finally:
            if excel:
                excel.Quit()
            pythoncom.CoUninitialize()

    # ------------------------------------------------------------------ openpyxl logic (fallback)
    def _run_logic_openpyxl(self, params_df, df, create_sheet, progress_callback):
        """Fallback sans COM — fonctionne en session 0/service."""
        results = []
        total_files = len(params_df)
        col_name = self.selected_column_name.strip().lower()

        for i, row in params_df.iterrows():
            file_path = str(row.get(col_name, "")).strip()
            logs = []

            file_params = {
                "file_path": file_path,
                "sheet_name": str(row.get("sheet_name", "Sheet1")).strip(),
                "start_col": self.safe_int(row.get("col", 1), 1),
                "start_row": self.safe_int(row.get("row", 1), 1)
            }

            try:
                abs_path = os.path.abspath(file_params["file_path"])

                if not os.path.isfile(abs_path):
                    if self.create_file_if_missing:
                        os.makedirs(os.path.dirname(abs_path), exist_ok=True)
                        logs.append("Fichier Excel introuvable, création automatique.")
                        wb = Workbook()
                        wb.save(abs_path)
                    else:
                        results.append({"file": file_path, "status": "ko", "details": "Fichier introuvable"})
                        continue

                ext = os.path.splitext(abs_path)[1].lower()
                keep_vba = ext in (".xlsm", ".xltm")
                wb = openpyxl.load_workbook(abs_path, keep_vba=keep_vba)

                sheet_name = file_params["sheet_name"]
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                else:
                    logs.append(f"Feuille '{sheet_name}' introuvable.")
                    if create_sheet:
                        ws = wb.create_sheet(title=sheet_name)
                        logs.append(f"Feuille '{sheet_name}' créée.")
                    elif "Sheet1" in wb.sheetnames:
                        ws = wb["Sheet1"]
                        logs.append("Fallback sur 'Sheet1'.")
                    else:
                        ws = wb.worksheets[0]
                        logs.append(f"Fallback sur '{ws.title}'.")

                wb.save(abs_path)

                detail_msg = "Insertion réussie (openpyxl)"
                if logs:
                    detail_msg += " (Note: " + " | ".join(logs) + ")"
                results.append({"file": file_path, "status": "ok", "details": detail_msg})

            except Exception as e:
                results.append({"file": file_params.get("file_path", "unknown"), "status": "ko", "details": str(e)})

            progress_callback(int(((i + 1) / total_files) * 100))

        return results

    def _run_logic(self, params_df, df, create_sheet, include_headers, progress_callback):
        """Détecte la session Windows et choisit la bonne méthode."""
        session = os.environ.get("SESSIONNAME", "")
        # Session 0 = service/planificateur sans desktop
        # Session Console ou RDP = session interactive
        if session.strip() == "" or session.strip().lower() == "services":
            print(f"[Excel] Session non-interactive ({session!r}), utilisation openpyxl")
            return self._run_logic_openpyxl(params_df, df, create_sheet, progress_callback)
        else:
            print(f"[Excel] Session interactive ({session!r}), utilisation COM")
            try:
                return self._run_logic_com(params_df, df, create_sheet, include_headers, progress_callback)
            except Exception as e:
                print(f"[Excel] COM échoué ({e}), fallback openpyxl...")
                return self._run_logic_openpyxl(params_df, df, create_sheet, progress_callback)

    # --- Helpers ---
    def safe_int(self, value, default=0):
        try:
            if value is None:
                return default

            value = str(value).strip()

            if value == "" or value == "?" or value.lower() == "nan":
                return default

            return int(float(value))
        except:
            return default

    def _send_status(self, status, details):
        domain = Domain([], metas=[DiscreteVariable("status", values=["ok", "ko"]), StringVariable("details")])
        self.Outputs.status_data.send(Table.from_list(domain, [[status, details]]))

    def _orange_table_to_dataframe(self, table: Table) -> pd.DataFrame:
        # Preserve original column order: attributes first, then metas, then class
        # but respect the domain's actual declaration order
        all_vars = list(table.domain.attributes) + list(table.domain.metas)
        if table.domain.class_var:
            all_vars.append(table.domain.class_var)

        data = {}
        for var in all_vars:
            raw = list(table.get_column(var.name))
            if isinstance(var, DiscreteVariable):
                data[var.name] = [
                    var.values[int(v)] if not (isinstance(v, float) and pd.isna(v)) and int(v) >= 0 else ""
                    for v in raw
                ]
            else:
                data[var.name] = raw

        df = pd.DataFrame(data).astype(str).replace("nan", "")
        
        # Re-order columns to match the original Table column order
        original_order = [var.name for var in table.domain.variables + table.domain.metas]
        df = df[[col for col in original_order if col in df.columns]]
        
        return df

if __name__ == "__main__":
    app = QApplication(sys.argv)
    w = OWInsertDataToExcel()
    w.show()
    if hasattr(app, "exec"):
        app.exec()
    else:
        app.exec_()
